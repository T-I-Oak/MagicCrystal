import openpyxl
import sys

def extract_xlsx_to_file(subfile, outfile):
    try:
        wb = openpyxl.load_workbook(subfile, data_only=True)
        with open(outfile, 'w', encoding='utf-8') as f:
            f.write(f"Sheets: {wb.sheetnames}\n\n")
            
            for sheetname in wb.sheetnames:
                f.write(f"--- Sheet: {sheetname} ---\n")
                ws = wb[sheetname]
                for i, row in enumerate(ws.iter_rows(values_only=True)):
                    # Replace None with empty string and join with pipe
                    row_str = [str(cell) if cell is not None else "" for cell in row]
                    if any(row_str):
                        # Add line number for reference
                        f.write(f"Row {i+1}: " + " | ".join(row_str) + "\n")
                f.write("\n")
        print(f"Successfully extracted to {outfile}")
    except Exception as e:
        print(f"Error: {e}")

if __name__ == "__main__":
    extract_xlsx_to_file("catfish.xlsx", "source_dump.txt")
